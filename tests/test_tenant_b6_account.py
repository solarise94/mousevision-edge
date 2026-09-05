"""B6 批次测试：登录自动激活 + 主账号 account 级汇总/跨工作区导出。

合同 docs/UPGRADE_TENANT_ISOLATION.md §4.2 / §15-B6：
- 登录自动激活：0 个可访问租户 → 不激活；恰好 1 个 → 自动 active_tenant；
  多个 → 不自动选（platform_admin 一律不激活，见 ui.app._auto_activate_tenant）。
- GET /api/account/summary：parent 看自己 A/B 不含未绑定 C；子账号 403。
- GET /api/account/export：跨工作区导出行带 tenant_id/tenant_name；子账号 403。

种子记录经设备凭证走 /api/records/report JSON 通道（B5），等价真实上报链路，
registry / records_meta / upload_queue / boxes 全部就位。
"""

from __future__ import annotations

import io
import json

import pytest
from fastapi.testclient import TestClient

from tenant_fixture import (  # noqa: F401 - pytest fixture 注册
    OPERATOR_PW,
    PLATFORM_ADMIN_PW,
    TENANT_ADMIN_PW,
    world,
)


# ------------------------------------------------------------------ #
# 种子工具
# ------------------------------------------------------------------ #
def _device_client(app) -> TestClient:
    """无 Cookie 的裸客户端：设备凭证是唯一凭证（会话优先级不干扰）。"""
    return TestClient(app)


def _report_record(world, slug: str, record_id: str, weight: float, cage: str) -> None:
    """以绑定该租户的设备凭证上报一条记录（JSON 通道，照片走占位兜底）。"""
    c = _device_client(world.app)
    r = c.post(
        "/api/records/report",
        json={
            "cage_id": cage,
            "project_id": "default",
            "records": [{"record_id": record_id, "ordinal": 1, "weight_g": weight}],
        },
        headers={"Authorization": f"Bearer {world.device_token(slug)}"},
    )
    assert r.status_code == 201, r.text
    assert r.json()["count"] == 1


@pytest.fixture()
def seeded(world):
    """a1 两条记录、a2 一条、b1 一条（各自独立箱号）。

    夹具 world 只给 a1/b1 签了设备，这里为 a2 补签一台（平台通道）。
    """
    r = world.platform.post(
        f"/api/control/tenants/{world.tid('a2')}/devices",
        json={"device_label": "phone-a2"},
    )
    assert r.status_code in (200, 201), r.text
    world.devices["a2"] = r.json()
    _report_record(world, "a1", "rec-a1-1", 11.11, "C57-101")
    _report_record(world, "a1", "rec-a1-2", 12.22, "C57-101")
    _report_record(world, "a2", "rec-a2-1", 22.22, "C57-202")
    _report_record(world, "b1", "rec-b1-1", 33.33, "C57-301")
    return world


# ------------------------------------------------------------------ #
# 登录自动激活（三情形 + 平台豁免 + 改密连续性）
# ------------------------------------------------------------------ #
def test_login_single_tenant_member_auto_activates(world):
    """单租户成员登录 → 会话自动 active_tenant，业务 API 无需显式激活。"""
    c = TestClient(world.app)
    r = c.post("/api/login", json={"username": "op-a1", "password": OPERATOR_PW})
    assert r.status_code == 200
    body = r.json()
    assert body["active_tenant_id"] == world.tid("a1")
    # 会话端点一致
    s = c.get("/api/control/session").json()
    assert s["active_tenant_id"] == world.tid("a1")
    assert s["roles"] == ["operator"]
    # 业务读/写无需再激活（保旧 UX）
    assert c.get("/api/overview").status_code == 200
    assert c.get("/api/records").status_code == 200


def test_login_multi_tenant_parent_not_auto_activated(world):
    """parent_owner 名下两个租户 → 不自动选；业务 API 403 直到显式切换。"""
    c = TestClient(world.app)
    r = c.post("/api/login", json={"username": "parent-a", "password": "ctl-parent-owner-pw"})
    assert r.status_code == 200
    body = r.json()
    assert body["active_tenant_id"] is None
    assert {t["tenant_id"] for t in body["tenants"]} == {world.tid("a1"), world.tid("a2")}
    assert c.get("/api/overview").status_code == 403
    # 显式激活后恢复
    r = c.post("/api/control/session/tenant", json={"tenant_id": world.tid("a1")})
    assert r.status_code == 200
    assert c.get("/api/overview").status_code == 200


def test_login_zero_tenant_user_not_activated(world):
    """零可访问租户（空 account 的 parent_owner）→ 不激活。"""
    r = world.platform.post(
        "/api/control/accounts",
        json={"name": "Lab C", "owner_username": "parent-c", "owner_password": "ctl-parent-c-pw"},
    )
    assert r.status_code in (200, 201), r.text
    c = TestClient(world.app)
    body = c.post(
        "/api/login", json={"username": "parent-c", "password": "ctl-parent-c-pw"}
    ).json()
    assert body["tenants"] == []
    assert body["active_tenant_id"] is None
    assert c.get("/api/overview").status_code == 403


def test_login_platform_admin_never_auto_activates(world):
    """platform_admin（即使恰好只有一个可访问租户）→ 一律不自动激活。

    保住 B3+B4 决策#1：账号级会话 + 显式设备/legacy 令牌 → 按令牌解析。
    """
    c = TestClient(world.app)
    # world 夹具已把 seed admin 改密为 PLATFORM_ADMIN_PW + "-changed"
    body = c.post(
        "/api/login",
        json={"username": "admin", "password": PLATFORM_ADMIN_PW + "-changed"},
    ).json()
    assert body["active_tenant_id"] is None
    # 激活的平台会话会压过设备凭证：这里只锁「未激活」这一前提
    assert c.get("/api/overview").status_code == 403


def test_password_change_new_session_stays_auto_activated(world):
    """改密撤销全部会话后补发的新会话仍自动激活（租户上下文连续）。"""
    c = TestClient(world.app)
    assert c.post(
        "/api/login", json={"username": "op-a1", "password": OPERATOR_PW}
    ).status_code == 200
    r = c.post(
        "/api/me/password",
        json={"current_password": OPERATOR_PW, "new_password": "ctl-operator-pw-2"},
    )
    assert r.status_code == 200
    s = c.get("/api/control/session").json()
    assert s["active_tenant_id"] == world.tid("a1")
    assert c.get("/api/overview").status_code == 200


def test_login_paused_tenant_not_auto_activated(world):
    """唯一可访问租户被 paused → 不自动激活（解析层拒绝其上下文）。"""
    world.control.set_tenant_status(world.tid("a1"), "paused")
    c = TestClient(world.app)
    body = c.post(
        "/api/login", json={"username": "op-a1", "password": OPERATOR_PW}
    ).json()
    assert body["active_tenant_id"] is None


def test_parent_owner_active_tenant_cannot_write_boxes(world):
    """Review 补缺口（API 级断言）：parent_owner 激活自己的租户后仍是只读
    作用域——写端点 POST /api/boxes → 403（§4.2，parent_owner 不在写角色）。"""
    parent = world.parent_client()
    r = parent.post("/api/control/session/tenant", json={"tenant_id": world.tid("a1")})
    assert r.status_code == 200, r.text
    assert set(r.json()["roles"]) == {"parent_owner"}

    r = parent.post(
        "/api/boxes", json={"cage_id": "C57-PARENT-W", "strain": "C57BL/6"}
    )
    assert r.status_code == 403, (
        f"parent_owner 激活租户后写端点必须 403（实际 {r.status_code}）: {r.text}"
    )
    # 写确实未发生：箱未创建；读通道（overview）仍可用（只读语义）
    assert parent.get("/api/boxes/C57-PARENT-W").status_code == 404
    assert parent.get("/api/overview").status_code == 200


# ------------------------------------------------------------------ #
# account 级汇总
# ------------------------------------------------------------------ #
def test_summary_parent_sees_own_tenants_not_others(seeded):
    c = seeded.parent_client()
    r = c.get("/api/account/summary")
    assert r.status_code == 200, r.text
    items = r.json()["items"]
    ids = {row["tenant_id"] for row in items}
    assert ids == {seeded.tid("a1"), seeded.tid("a2")}
    names = {row["tenant_name"] for row in items}
    assert names == {"Workspace A1", "Workspace A2"}
    # 未绑定的 account B 工作区绝不出现
    assert seeded.tid("b1") not in ids
    assert "Workspace B1" not in json.dumps(r.json(), ensure_ascii=False)


def test_summary_platform_sees_all_active(seeded):
    c = TestClient(seeded.app)
    r = c.post(
        "/api/login",
        json={"username": "admin", "password": PLATFORM_ADMIN_PW + "-changed"},
    )
    assert r.status_code == 200, r.text
    r = c.get("/api/account/summary")
    assert r.status_code == 200
    ids = {row["tenant_id"] for row in r.json()["items"]}
    assert {seeded.tid("a1"), seeded.tid("a2"), seeded.tid("b1")} <= ids


def test_summary_row_shape_and_counts(seeded):
    c = seeded.parent_client()
    items = c.get("/api/account/summary").json()["items"]
    by_id = {row["tenant_id"]: row for row in items}
    a1 = by_id[seeded.tid("a1")]
    for key in (
        "tenant_id",
        "tenant_name",
        "account_id",
        "account_name",
        "status",
        "boxes",
        "records",
        "pending_uploads",
        "last_sync_at",
    ):
        assert key in a1, f"缺字段 {key}"
    assert a1["status"] == "active"
    assert a1["boxes"] == 1  # C57-101
    assert a1["records"] == 2  # 上报即 records_meta/队列 Pending
    assert a1["pending_uploads"] == 2
    assert a1["last_sync_at"]  # 设备上报已触碰 last_used_at
    a2 = by_id[seeded.tid("a2")]
    assert a2["records"] == 1
    assert a2["pending_uploads"] == 1


def test_summary_child_member_403(seeded):
    c = seeded.member_client("child", "op-a1", OPERATOR_PW, "a1")
    assert c.get("/api/account/summary").status_code == 403


def test_summary_tenant_admin_403(seeded):
    c = seeded.member_client("child-admin", "admin-a1", TENANT_ADMIN_PW, "a1")
    assert c.get("/api/account/summary").status_code == 403


def test_summary_anonymous_401(seeded):
    assert TestClient(seeded.app).get("/api/account/summary").status_code == 401


# ------------------------------------------------------------------ #
# 跨工作区导出
# ------------------------------------------------------------------ #
def test_export_csv_has_tenant_columns_and_scope(seeded):
    c = seeded.parent_client()
    r = c.get("/api/account/export", params={"format": "csv"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    text = r.content.decode("utf-8-sig")
    header = text.splitlines()[0]
    assert header.split(",")[:2] == ["tenant_id", "tenant_name"]
    # A1/A2 的数据在，B1 的不在
    assert seeded.tid("a1") in text and seeded.tid("a2") in text
    assert "11.11" in text and "22.22" in text
    assert seeded.tid("b1") not in text
    assert "33.33" not in text


def test_export_csv_scope_matches_summary_for_parent(seeded):
    c = seeded.parent_client()
    text = c.get("/api/account/export", params={"format": "csv"}).content.decode("utf-8-sig")
    rows = [line for line in text.splitlines()[1:] if line.strip()]
    assert len(rows) == 3  # a1 两条 + a2 一条


def test_export_xlsx_loadable_with_tenant_columns(seeded):
    from openpyxl import load_workbook

    c = seeded.parent_client()
    r = c.get("/api/account/export", params={"format": "xlsx"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["records"]
    header = [cell.value for cell in ws[1]]
    assert header[:2] == ["tenant_id", "tenant_name"]
    body = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    assert len(body) == 3
    tids = {row[0] for row in body}
    assert tids == {seeded.tid("a1"), seeded.tid("a2")}


def test_export_child_member_403(seeded):
    c = seeded.member_client("child", "op-a1", OPERATOR_PW, "a1")
    r = c.get("/api/account/export", params={"format": "csv"})
    assert r.status_code == 403


def test_export_anonymous_401(seeded):
    r = TestClient(seeded.app).get("/api/account/export", params={"format": "csv"})
    assert r.status_code == 401


def test_export_bad_format_422(seeded):
    c = seeded.parent_client()
    assert c.get("/api/account/export", params={"format": "pdf"}).status_code == 422
